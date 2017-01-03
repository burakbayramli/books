from __future__ import division
from __future__ import print_function
from numpy import array
from pandas import read_csv
from pandas import read_excel
from pandas import read_stata
import io
import openpyxl
import scipy.io as sio
import sys
import xlrd
from pylab import *
from numpy import *
# End Imports


csv_data = read_csv('FTSE_1984_2012.csv')
csv_data = csv_data.values
csv_data[:4]
open = csv_data[:,1]

csv_data = read_csv('FTSE_1984_2012_numeric.csv')
csv_data = csv_data.values
csv_data[:4,:2]

excel_data = read_excel('FTSE_1984_2012.xls','FTSE_1984_2012')
excel_data = excel_data.values
excel_data[:4,:2]
open = excel_data[:,1]

stata_data = read_stata('FTSE_1984_2012.dta')
stata_data = stata_data.values
stata_data[:4,:2]

try:
    data = loadtxt('FTSE_1984_2012.csv',delimiter=',') # Error
except:
    print("Error detected in: data = loadtxt('FTSE_1984_2012.csv',delimiter=',') # Error")
    error = sys.exc_info()
    print("Error type: " + str(error[0]) + " Error message: " + str(error[1]))
try:
    data = loadtxt('FTSE_1984_2012_numeric.csv',delimiter=',') # Error
except:
    print("Error detected in: data = loadtxt('FTSE_1984_2012_numeric.csv',delimiter=',') # Error")
    error = sys.exc_info()
    print("Error type: " + str(error[0]) + " Error message: " + str(error[1]))
data = loadtxt('FTSE_1984_2012_numeric.csv',delimiter=',',skiprows=1)
data[0]

data = genfromtxt('FTSE_1984_2012.csv',delimiter=',')
data[0]
data[1]

data = genfromtxt('FTSE_1984_2012_numeric_tab.txt',delimiter='\t')

data = csv2rec('FTSE_1984_2012.csv',delimiter=',')
data[0]

open = data['open']
open

wb = xlrd.open_workbook('FTSE_1984_2012.xls')
# To read xlsx change the filename
# wb = xlrd.open_workbook('FTSE_1984_2012.xlsx')
sheetNames = wb.sheet_names()
# Assumes 1 sheet name
sheet = wb.sheet_by_name(sheetNames[0])
excelData = [] # List to hold data
for i in xrange(sheet.nrows):
    excelData.append(sheet.row_values(i))
# Subtract 1 since excelData has the header row
open = empty(len(excelData) - 1)
for i in xrange(len(excelData) - 1):
    open[i] = excelData[i+1][1]

wb = openpyxl.load_workbook('FTSE_1984_2012.xlsx')
sheetNames = wb.get_sheet_names()
# Assumes 1 sheet name
sheet = wb.get_sheet_by_name(sheetNames[0])
rows = sheet.rows
# Subtract 1 since excelData has the header row
open = empty(len(rows) - 1)
for i in xrange(len(excelData) - 1):
    open[i] = rows[i+1][1].value

wb = openpyxl.load_workbook('FTSE_1984_2012.xlsx', use_iterators = True)
sheetNames = wb.get_sheet_names()
# Assumes 1 sheet name
sheet = wb.get_sheet_by_name(sheetNames[0])
# Use list to store data
open = []
# Changes since access is via memory efficient iterator
# Note () on iter_rows
for row in sheet.iter_rows():
    # Must use internal_value
    open.append(row[1].internal_value)
# Burn first row and convert to array
open = array(open[1:])

matData = sio.loadmat('FTSE_1984_2012.mat')
type(matData)
matData.keys()
open = matData['open']

f = io.open('IBM_TAQ.txt', 'r')
line = f.readline()
# Burn the first list as a header
line = f.readline()
date = []
time = []
price = []
volume = []
while line:
    data = line.split(',')
    date.append(int(data[1]))
    price.append(float(data[3]))
    volume.append(int(data[4]))
    t = data[2]
    time.append(int(t.replace(':','')))
    line = f.readline()
# Convert to arrays, which are more useful than lists
# for numeric data
date = array(date)
price = array(price)
volume = array(volume)
time = array(time)
allData = array([date,price,volume,time])
f.close()

x = arange(10)
y = zeros((100,100))
savez_compressed('test',x,y)
data = load('test.npz')
# If no name is given, arrays are generic names arr_1, arr_2, etc
x = data['arr_1']
savez_compressed('test',x=x,otherData=y)
data = load('test.npz')
# x=x provides the name x for the data in x
x = data['x']
# otherDate = y saves the data in y as otherData
y = data['otherData']

x = array([1.0,2.0,3.0])
y = zeros((10,10))
# Set up the dictionary
saveData = {'x':x, 'y':y}
sio.savemat('test',saveData,do_compression=True)
# Read the data back in
matData = sio.loadmat('test.mat')

x = randn(10,10)
# Save using tabs
savetxt('tabs.txt',x)
# Save to CSV
savetxt('commas.csv',x,delimiter=',')
# Reread the data
xData = loadtxt('commas.csv',delimiter=',')

