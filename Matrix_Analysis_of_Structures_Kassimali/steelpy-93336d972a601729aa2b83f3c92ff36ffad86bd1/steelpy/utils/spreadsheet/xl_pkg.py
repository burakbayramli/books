# Copyright (c) 2022 steelpy
#
# Python stdlib imports
from __future__ import annotations

# from dataclasses import dataclass
from pathlib import Path
import re

# package imports
from steelpy.utils.dataframe.main import DBframework
from steelpy.utils.spreadsheet.pylightxl.pylightxl import(utility_address2index,
                                                            utility_index2address)
from .xlops import get_cell_name, get_cellcol_name


#
class ExcelExt:
    __slots__ = ['units', '_wb', '_ws', '_flag']

    def __init__(self, wb_name: str):
        """
        """
        self._flag = True
        try:
            import xlwings as xw
            xw.Book(wb_name).set_mock_caller()
            self._wb = xw.Book.caller()
            self._flag = False
        except ModuleNotFoundError:
            try:
                from openpyxl import load_workbook
                #self._wb = load_workbook(wb_name, read_only=False,
                #                         keep_vba=True, data_only=True)
                self._wb = load_workbook(wb_name)
            except PermissionError:
                raise PermissionError('close woorkbook or install xlwings')
        #
        #
        self._ws = Sheets(self._wb)
        # print('-->')

    #
    @property
    def sheets(self):
        """ """
        return self._ws    
    #
    #
    #
    @property
    def sheet_names(self):
        """ """
        return self._ws.ws_names  
    #
    #
    def save(self):
        """ """
        if self._flag:
            self._ws.close()


#
#
class Sheets:
    __slots__ = ['units', '_wb', '_cells']

    def __init__(self, wb: str):
        """
        """
        self._wb = wb
    #
    def __getitem__(self, ws_name: str):
        """
        """
        s_name = self.ws_names
        if ws_name in s_name:
            return SheetOps(self._wb, ws_name)
        else:
            raise IOError(f"sheet {ws_name} not found")
    #
    @property
    def ws_names(self):
        """ get sheet names"""
        try:
            return [sh.name for sh in self._wb.sheets]
        except AttributeError:
            return self._wb.sheetnames
    #
    def close(self):
        """ """
        #if self._flag:
        self._wb.close()    
#
#
#
class SheetOps:
    __slots__ = ['_wb', '_ws_name', '_ws', '_df']

    def __init__(self, wb, ws_name: str):
        """ ws : sheet"""
        self._wb = wb
        self._ws_name = ws_name
        try: # xlwings
            self._ws = self._wb.sheets[self._ws_name]
        except AttributeError: # openpyxl
            self._ws  = self._wb[self._ws_name]
        #
        self._df = DataFrameOps(self._ws)
    #
    #
    def __setitem__(self, cell_name: str, value) -> None:
        """
        """
        # if re.search(r"\:", cell_name):
        # tokens = re.split(r'[:]', cell_name)
        # row_i, col_i = utility_address2index(tokens[0])
        # row_j, col_j = utility_address2index(tokens[1])
        # for i in range(row_i, row_j+1):
        #    for j in range(col_i, col_j+1):
        #        self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        if isinstance(value, list):
            row_i, col_j = utility_address2index(cell_name)
            columns = len(value)
            if isinstance(value[0], (list,tuple)):
                try: # xlwings
                    self._ws.range(cell_name).value = value
                except AttributeError: # openpyxl
                    rows =  len(value[0])
                    for j in range(columns):  # columns
                        for i in range(rows):  # rows
                            #print('---', i)
                            self._ws.cell(row=row_i + i,
                                          column=col_j + j,
                                          value=value[j][i])
            else:
                try: # xlwings
                    self._ws.range(cell_name).value = value
                except AttributeError: # openpyxl
                    for j in range(columns):  # columns
                        self._ws.cell(row=row_i,
                                      column=col_j + j,
                                      value=value[j])
        else:
            cell_name = get_cell_name(cell_name)
            try: # xlwings
                self._ws.range(cell_name).value = value
            except AttributeError: # openpyxl
                self._wb[cell_name] = value

    #
    def __getitem__(self, cell_name: str):
        """
        """
        try:
            #ws = self._wb.sheets[self._ws_name]
            return self._ws.range(cell_name).value
        except AttributeError:
            #ws = self._wb[self._ws_name ]
            #return [[item.value for item in rows]
            #        for rows in self._ws[cell_name]]
            return self._ws[cell_name].value
    #
    @property
    def column(self):
        """ """
        return xlColumn(self._ws)
    #
    @property
    def row(self):
        """ """
        return xlRow(self._ws)     
    #
    #
    def key(self, row:int|str|None=None,
            column:int|str|None=None):
        """ """
        #ws_name = self._ws_name
        #
        #if row:
        #    if isinstance(column, str):
        #        
        #
        #if column:
        #    if isinstance(column, str):
        return self._wb.ws(ws=self._ws_name).ssd(keyrows=row, keycols=column)
    #
    #
    #def to_df(self, row:int=1, column:int|str=1,
    #          skiprows:int|list=0, names:str|list|None=None):
    #    """ convert sheet data to dataframe"""
    #    return self._df.get_df(row, column, skiprows, names)
    #
    def to_df(self, skiprows:int|list=0, names:str|list|None=None):
        """ convert sheet data to dataframe"""
        return self._df.get_df(skiprows=skiprows, names=names)
    #
    #
    #@property
    def df_tool(self, index:int|bool=False, header:bool=False,
                dates:None=None):
        """ 
        index : When writing, include or exclude the index by setting it to True or False
        header : When writing, include or exclude the index and series names by setting it to True or False
        dates : 
        """
        return DataFrameOps(self._ws, index=index, header=header)
    #
    def get_data(self):
        """ """
        try:
            data = self._ws.used_range.value
        except AttributeError:
            data = get_column_data(self._ws)
        return data    
#
#
#
class DataFrameOps:
    __slots__ = ['_ws',  '_column', '_index', '_header', '_df']

    def __init__(self, ws, index:int|bool=False, header:bool=False):
        """ ws : sheet"""
        self._ws = ws
        self._index = index
        self._header = header
        self._column = xlColumn(self._ws)
        self._df = DBframework()
    #
    #
    def __setitem__(self, cell_name:str|int|tuple, df) -> None:
        """
        """
        cell_name = get_cell_name(cell_name)
        
        if re.search(r"\:", cell_name):
            1/0
        else:
            # [row, col]
            index = utility_address2index(cell_name)
            for x, (key, value) in enumerate(df.items()):
                row = index[0]
                col = index[1] + x
                address = utility_index2address(row, col)
                #self._ws.range(address).value = key
                if self._header:
                    try:
                        self._ws[address].value = key
                    except AttributeError:
                        self._ws[address] = key
                    row += 1
                #
                address = utility_index2address(row, col)
                self._column[address] = value.tolist()
    # __call__
    def get_df(self, skiprows:int|list=0,
               names:str|list|None=None):
        """ """
        #if not db_name:
        #    db_name = self._ws_name
        #
        data = self.get_data()
        columns = list(zip(*data))
        #
        if isinstance(skiprows, (list, tuple)):
            print('---')
            1/0
        else:
            if names:
                data = {names[idx]: col[skiprows+1:]
                        for idx, col in enumerate(columns)}
            else:
                data = {col[skiprows]: col[skiprows+1:]
                        for col in columns}
        #
        #print('--')
        return self._df.DataFrame(data, columns=names)
    #
    #
    def get_data(self):
        """ """
        try:
            data = self._ws.used_range.value
        except AttributeError:
            data = get_column_data(self._ws)
        return data     
#
#
class WriteExcelData:
    __slots__ = ['_wb', '_wb_name', '_flag',
                 '_col2name']

    def __init__(self, wb_name):
        """
        """
        self._flag = True
        try:
            import xlwings as xw
            xw.Book(wb_name).set_mock_caller()
            self._wb = xw.Book.caller()
            self._flag = False
        except ModuleNotFoundError:
            from openpyxl import load_workbook
            self._wb = load_workbook(wb_name)
        #
        path = Path().absolute()
        path = str(path) + f"\\{wb_name}"
        self._wb_name = path
        #
        self._col2name = ColumnName()
    #
    #
    def sheet_active(self, ws_name: str):
        """ """
        try:
            return self._wb[ws_name]
        except TypeError:
            # return self._wb.sheets[ws_name]
            return xlset(ws=self._wb.sheets[ws_name])

    #
    #
    def cell_name(self, column: int, row: int):
        """ return cell name"""
        col = self._col2name.get_column_name(column)
        name = f"{col}{row}".upper()
        return name

    #
    #
    # def cell(self,row:int, column:int, value:Union[str,int,float]):
    #    """ """
    #
    #
    def plot(self, ws_name: str, fig, anchor: str, name: str):
        """ """
        try:
            sheet = self._wb[ws_name]
            from openpyxl.drawing.image import Image
            fig.savefig('unit.png')
            img = Image('unit.png')
            # img.width = 640
            # img.height = 780
            img.anchor = anchor  # "I50"
            sheet.add_image(img)
        except TypeError:
            sheet = self._wb.sheets[ws_name]
            sheet.pictures.add(fig, name=name, update=True)
            #

    def update_column(self, ws_name: str, col: int,
                      values: dict, step: int):
        """ """
        try:  # openpyxl
            cells = self._wb[ws_name]
        except TypeError:  # xlwings
            ws = self._wb.sheets[ws_name]
            # cells = ws.cells
            cells = ws.range((1, 1), (135, 16))
        #
        for row in cells.rows:
            name = row[col].value
            try:
                for x, item in enumerate(values[name]):
                    row[step + x].value = item
            except KeyError:
                continue
        # print('--')

    #
    def close(self):
        """ """
        if self._flag:
            self._wb.save(self._wb_name)
        #


#
class NewExcelFileXX:
    __slots__ = ['_wb', '_wb_name']

    def __init__(self, wb_name: str):
        """
        """
        self._wb = Database()
        self._wb_name = wb_name
    #
    #
    def sheet_active(self, ws_name: str):
        """ """
        self._wb.add_ws(ws=ws_name)
        return xlsetNew(wb=self._wb, ws_name=ws_name)

    #
    def close(self, file_name):
        """ """
        #fn = self._wb_name + ".xlsx"
        #fn = os.path.join(self._path, fn)
        writexl(db=self._wb, fn=file_name)


#
class xlsetNew:
    __slots__ = ['_wb', '_ws_name']

    def __init__(self, wb, ws_name: str):
        """ ws : sheet"""
        self._wb = wb
        self._ws_name = ws_name

    #
    #
    def __setitem__(self, cell_name: str, value) -> None:
        """
        """
        # if re.search(r"\:", cell_name):
        # tokens = re.split(r'[:]', cell_name)
        # row_i, col_i = utility_address2index(tokens[0])
        # row_j, col_j = utility_address2index(tokens[1])
        # for i in range(row_i, row_j+1):
        #    for j in range(col_i, col_j+1):
        #        self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        if isinstance(value, list):
            row_i, col_j = utility_address2index(cell_name)
            columns = len(value)
            if isinstance(value[0], list):
                rows =  len(value[0])
                for j in range(columns):  # columns
                    for i in range(rows):  # rows
                        print('---', i)
                        self._wb.ws(ws=self._ws_name).update_index(row=row_i + i, 
                                                                   col=col_j + j,
                                                                   val=value[j][i])                        
            else:
                for j in range(columns):  # columns
                    self._wb.ws(ws=self._ws_name).update_index(row=row_i, 
                                                               col=col_j + j,
                                                               val=value[j])
        else:
            self._wb.ws(ws=self._ws_name).update_address(address=cell_name, val=value)

    #
    def __getitem__(self, cell_name: str):
        """
        """
        if re.search(r"\:", cell_name):
            tokens = re.split(r'[:]', cell_name)
            row_i, col_i = utility_address2index(tokens[0])
            row_j, col_j = utility_address2index(tokens[1])
            for i in range(row_i, row_j + 1):
                for j in range(col_i, col_j + 1):
                    self._wb.ws(ws=self._ws_name).update_index(row=i, col=j, val=value)
        else:
            return self._wb.ws(ws=self._ws_name).index(address=cell_name)
            # row_id, col_id = utility_address2index(cell_name)
            # return self._wb.ws(ws=self._ws_name).index(row=row_id, col=col_id)
#
#
def get_column_data(sheet):
    """ """
    cell_obj = []
    for row in sheet.rows:
        cell_obj.append([])
        for cell in row:
            cell_obj[-1].append(cell.value)
    return cell_obj
#
#
class xlset:
    __slots__ = ['_ws']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws

    #
    #
    def __setitem__(self, cell_name: str, value: float) -> None:
        """
        """
        self._ws.range(cell_name).value = value

    #
    def __getitem__(self, cell_name: int):
        """
        """
        return self._ws.range(cell_name).value
#
#
def get_row_column(data_list: list, header: str):
    """ """
    row = [x for x, item in enumerate(data_list) if header in item]
    col = [x for x, item in enumerate(data_list[row[0]]) if item == header]
    return [row[0], col[0]]
#
#
#
class xlColumn:
    __slots__ = ['_ws']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws
        #self._ws_name = ws_name
    #
    def __setitem__(self, col_name:int|str|tuple, 
                    value:float|int|str|tuple|list) -> None:
        """
        """
        cname = get_cellcol_name(col_name)
        #
        # xlwings
        try:
            self._ws.range(cname).options(transpose=True).value = value
        except: # openpyxl
            #ws = self._wb[self._ws_name]
            row, col = utility_address2index(cname)
            for x, item in enumerate(value):
                row += x
                self._ws.cell(row=row, column=col).value = item         
    #
    #
    def __getitem__(self, col_name: int|str):
        """
        """
        if isinstance(col_name, int):
            cname = f"{col_name}1"
        
        elif isinstance(col_name, str):
            numbers = re.findall('[0-9]+', col_name)
            if numbers:
                cname = col_name
            else:
                cname = f"{col_name}1"
        
        else:
            raise IOError(f"column name : {col_name} not valid")
        #
        #
        try: # xlwings
            return self._ws.range(cname).expand("down").value

        except AttributeError: # openpyxl
            #ws = self._wb[self._ws_name]
            row, col = utility_address2index(cname)
            cols = tuple(self._ws.columns)
            data = [item.value for item in cols[col-1]]
            return data[row-1:]
            #try:
            #    idx = data.index(None)
            #except ValueError:
            #    idx = -1
            #return data[:idx]    
#
class xlRow:
    __slots__ = ['_ws']

    def __init__(self, ws):
        """ ws : sheet"""
        self._ws = ws
    #
    def __setitem__(self, row_name: str, value: float) -> None:
        """
        """
        if isinstance(row_name, int):
            rname = f"A{row_name}"

        elif isinstance(row_name, str):
            numbers = re.findall('[0-9]+', row_name)
            if numbers:
                rname = row_name
            else:
                rname = f"A{row_name}"
        
        else:
            raise IOError(f"row name : {row_name} not valid")         
        #
        try:  # xlwings
            self._ws.range(rname).expand("right").value = value
        
        except AttributeError:  # openpyxl
            #ws = self._wb[self._ws_name]
            row, col = utility_address2index(rname)
            for x, item in enumerate(value):
                col += x
                self._ws.cell(row=row, column=col).value = item
    #
    def __getitem__(self, row_name:int|str):
        """
        """
        if isinstance(row_name, int):
            rname = f"A{row_name}"

        elif isinstance(row_name, str):
            numbers = re.findall('[0-9]+', row_name)
            if numbers:
                rname = row_name
            else:
                rname = f"A{row_name}"
        
        else:
            raise IOError(f"row name : {row_name} not valid")
        #
        try:  # xlwings
            return self._ws.range(rname).expand("right").value
        
        except AttributeError:  # openpyxl
            #ws = self._wb[self._ws_name]
            row, col = utility_address2index(rname)
            rows = tuple(self._ws.rows)
            data = [item.value for item in rows[row-1]]
            return data[col-1:]
            #try:
            #    idx = data.index(None)
            #except ValueError:
            #    idx = -1
            #return data[:idx]


#
#
#
